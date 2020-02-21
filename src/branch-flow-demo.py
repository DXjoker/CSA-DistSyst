# !/usr/bin/python
#
# Copyright 2020, Southeast University, Liu Pengxiang
#
# IEEE PES-GM 2020 Paper
# Demo of Branch-flow model (also known as DistFlow model)


import math
import numpy as np
import gurobipy as gp
import openpyxl as pyxl
import matplotlib.pyplot as plt


# This class creates the parameter class
# -----------------------------------------------------------------------------
#
class Parameter(object):

    # Initialization
    def __init__(self, filename):

        # 1. System Data
        tool = Excel_tool()
        data = tool.read_excel(filename)
        # 1) line
        self.Line   = data[0]
        self.N_line = len(self.Line) # number of line
        # 2) bus
        self.Bus    = data[1]
        self.N_bus  = len(self.Bus)  # number of bus
        # 3) substation
        self.Sub    = data[2]
        self.N_sub  = len(self.Sub)  # number of substation
        # 4) generator
        self.Gen    = data[3]
        self.N_gen  = len(self.Gen)  # number of renewables
        self.Factor = 0.31756  # power factor (rad)
        # 5) daily curve
        self.Day    = data[4]
        self.N_time = len(self.Day)  # number of hours
        # 6) global variable index
        self.var_index()

        # Base value
        self.Base_V = 12.66  # voltage: 12.66 kV
        self.Base_S = 10.00  # power:   10.00 MVA
        self.Base_Z = self.Base_V ** 2 / self.Base_S  # impedance
        self.Base_I = self.Base_S / self.Base_V / np.sqrt(3)  # current
        # Cost
        self.Cost_sub = 83   # cost of power purchasing
        self.Cost_pen = 200  # cost of load shedding
        self.Cost_los = 25   # cost of power loss
        # Other
        self.Big_M = 1e2  # a sufficient large number
        self.V_min = (0.95 * self.Base_V) ** 2
        self.V_max = (1.05 * self.Base_V) ** 2
        # Bus-Line Information
        self.Line_head = [[] for i in range(self.N_bus)]
        self.Line_tail = [[] for i in range(self.N_bus)]
        for i in range(self.N_line):
            head = self.Line[i][1]
            tail = self.Line[i][2]
            self.Line_head[int(round(head))].append(i)
            self.Line_tail[int(round(tail))].append(i)
    
    # This function creates global index
    def var_index(self):

        # Fictitious power flow
        # 1. Name
        global N_F_line, N_F_sub , N_F_load, N_F_gen , N_F_var
        # 2. Initialization
        N_F_line = 0                       # flow of line
        N_F_load = N_F_line + self.N_line  # flow of load demand
        N_F_sub  = N_F_load + self.N_bus   # flow of substation
        N_F_gen  = N_F_sub  + self.N_sub   # flow of DG
        N_F_var  = N_F_gen  + self.N_gen   # Number of all variables

        # Real power flow
        # 1. Name
        global N_V_bus, N_I_line, N_P_line, N_Q_line, N_P_sub, N_Q_sub
        global N_S_gen, N_C_gen , N_P_cut , N_Q_cut , N_N_var
        # 2. Initialization
        N_V_bus  = 0                       # square of voltage amplitude
        N_I_line = N_V_bus  + self.N_bus   # square of voltage phase angle
        N_P_line = N_I_line + self.N_line  # power flow (active)
        N_Q_line = N_P_line + self.N_line  # power flow (reactive)
        N_P_sub  = N_Q_line + self.N_line  # power injection at substation
        N_Q_sub  = N_P_sub  + self.N_sub   # power injection at substation
        N_S_gen  = N_Q_sub  + self.N_sub   # renewables generation
        N_C_gen  = N_S_gen  + self.N_gen   # renewables curtailment
        N_P_cut  = N_C_gen  + self.N_gen   # Load shedding (active)
        N_Q_cut  = N_P_cut  + self.N_bus   # Load shedding (reactive)
        N_N_var  = N_Q_cut  + self.N_bus   # Number of all variables


# This class creates the execution tool for Excel files
# -----------------------------------------------------------------------------
# 
class Excel_tool(object):

    # Initialization
    def __init__(self):
        pass

    # inputs data from Excel file
    def read_excel(self, filename):
        data = []
        book = pyxl.load_workbook(filename)
        # Data preprocessing
        for i, name in enumerate(book.sheetnames):  # sheet number
            if i < len(book.sheetnames):
                sheet = book[name]
                n_row = sheet.max_row     # number of rows
                n_col = sheet.max_column  # number of columns
                data.append(self.tool_filter(sheet, n_row, n_col))
        return data

    # saving data to excel file
    def save_excel(self, filename, sheetname, title, data):
        book  = pyxl.load_workbook(filename)
        name  = book.sheetnames[-1]
        book.remove(book[name])
        sheet = book.create_sheet(sheetname)
        sheet.append(title)
        for i in range(len(data)):
            sheet.append(data[i,:].tolist())  # write data
        book.save(filename = filename)
    
    # filter data from the original numpy array
    def tool_filter(self, sheet, n_row, n_col):
        k = 0
        data = []
        for i in range(n_row):
            if sheet['A' + str(i + 1)].data_type == 'n':  # if it is a number
                data.append([])
                for j in range(n_col):
                    pos = chr(64 + j + 1) + str(i + 1)  # the position
                    val = sheet[pos].value
                    if sheet[pos].data_type == 'n':
                        data[k].append(val)
                k = k + 1
        return np.array(data)


# This class creates the grapher class
# -----------------------------------------------------------------------------
#
class Grapher(object):

    # Initialization
    def __init__(self, Para):
        # Switch line coordinate
        self.switch_line = []
        self.switch_line.append(np.array([[ 7, 0], [ 7,-2], [3 ,-2], [ 3,-3]]))
        self.switch_line.append(np.array([[ 8, 0], [ 8, 2], [14, 2], [14, 0]]))
        self.switch_line.append(np.array([[11, 0], [11,-3], [ 4,-3]]))
        self.switch_line.append(np.array([[17, 0], [17, 3], [12, 3]]))
        self.switch_line.append(np.array([[ 8, 6], [ 8, 5], [ 8, 3]]))
        # Line coordinate
        self.coordinate = []
        for n in range(Para.N_line):
            if Para.Line[n,6] == 0:  # line
                bus_head = int(round(Para.Line[n,1]))
                bus_tail = int(round(Para.Line[n,2]))
                x0 = Para.Bus[bus_head, 3]
                y0 = Para.Bus[bus_head, 4]
                x1 = Para.Bus[bus_tail, 3]
                y1 = Para.Bus[bus_tail, 4]
                self.coordinate.append(np.array([[x0,y0], [x1, y1]]))
            else:
                switch_no = int(Para.Line[n,6] - 1)
                self.coordinate.append(self.switch_line[switch_no])
    
    # Figuring
    def figure(self, Para, sol):
        self.plot_node(Para)
        self.plot_line(Para, sol)
        plt.axis('equal')
        plt.show()

    # Node
    def plot_node(self, Para):
        for n in range(Para.N_bus):
            plt.plot(Para.Bus[n,3], Para.Bus[n,4], 'b.')
            plt.text(Para.Bus[n,3] + 0.05, Para.Bus[n,4] + 0.10, '%s' % n)
    
    # Line
    def plot_line(self, Para, sol):
        for n in range(Para.N_line):
            for m in range(np.size(self.coordinate[n], 0) - 1):
                x0 = self.coordinate[n][m, 0]
                y0 = self.coordinate[n][m, 1]
                x1 = self.coordinate[n][m + 1, 0]
                y1 = self.coordinate[n][m + 1, 1]
                if sol.y_line[n, 0] == 1:
                    plt.plot([x0, x1], [y0, y1], 'b-' )
                if sol.y_line[n, 0] == 0:
                    plt.plot([x0, x1], [y0, y1], 'r--')


# This class creates the Result class
# -----------------------------------------------------------------------------
# 
class Result(object):
    
    # Initialization
    def __init__(self):
        pass

    # Obtain the value of variables
    def get_value(self, model, var, var_type):
        # Get value
        key = var.keys()
        val = var.copy()
        for i in range(len(key)):
            val[key[i]] = var[key[i]].x
        # Calculate dimention
        if isinstance(max(key),tuple):  # multi dimention
            dim = tuple([item + 1 for item in max(key)])
        if isinstance(max(key),int):    # one   dimention
            dim = tuple([int(len(key)),1])
        # Convert dictionary to numpy array
        arr = np.zeros(dim, dtype = var_type)
        if var_type == "int":
            for i in range(len(val)):
                arr[key[i]] = int(round(val[key[i]]))
        else:
            for i in range(len(val)):
                arr[key[i]] = round(val[key[i]], 4)
        return arr


# This function defines the branch flow model
# -----------------------------------------------------------------------------
#
def main_function(Para, hour):
    
    # Initialization
    # Import gurobi model
    model = gp.Model()

    # Topology variables
    y_line = model.addVars(Para.N_line, vtype = "B")  # binary
    # Operating variable
    f_flow = model.addVars(N_F_var, lb = -1e2)  
    v_flow = model.addVars(N_N_var, lb = -1e2)
    # Set objective
    obj    = model.addVars(1)

    # Build Models
    # Build the model
    model = Reconfig_Model(model, Para, y_line, f_flow)
    model = DistFlow_Model(model, Para, y_line, v_flow, hour, obj)

    # Objective
    model.setObjective(obj.sum('*'), gp.GRB.MINIMIZE)

    # Optimization
    model.optimize()
    if model.status == gp.GRB.Status.OPTIMAL:
        sol = Result()
        sol.y_line = sol.get_value(model, y_line, "int")
        sol.f_flow = sol.get_value(model, f_flow, "float")
        sol.v_flow = sol.get_value(model, v_flow, "float")
        sol.V_bus  = sol.v_flow[N_V_bus  : N_V_bus  + Para.N_bus ]
        sol.I_line = sol.v_flow[N_I_line : N_I_line + Para.N_line]
        sol.P_line = sol.v_flow[N_P_line : N_P_line + Para.N_line]
        sol.Q_line = sol.v_flow[N_Q_line : N_Q_line + Para.N_line]
        sol.P_sub  = sol.v_flow[N_P_sub  : N_P_sub  + Para.N_sub ]
        sol.Q_sub  = sol.v_flow[N_Q_sub  : N_Q_sub  + Para.N_sub ]
        sol.S_gen  = sol.v_flow[N_S_gen  : N_S_gen  + Para.N_gen ]
        sol.C_gen  = sol.v_flow[N_C_gen  : N_C_gen  + Para.N_gen ]
        sol.P_cut  = sol.v_flow[N_P_cut  : N_P_cut  + Para.N_bus ]
        sol.Q_cut  = sol.v_flow[N_Q_cut  : N_Q_cut  + Para.N_bus ]
    else:
        sol = -1

    return sol


# This function defines the reconfiguration model
# -----------------------------------------------------------------------------
#
def Reconfig_Model(model, Para, y_line, f_flow):
    
    # Constraint

    # 1. Fictitious power flow
    for n in range(Para.N_line):
        model.addConstr(f_flow[N_F_line + n] >= -1e2 * y_line[n])
        model.addConstr(f_flow[N_F_line + n] <=  1e2 * y_line[n])
    for n in range(Para.N_sub):
        model.addConstr(f_flow[N_F_sub  + n] >=  0)
        model.addConstr(f_flow[N_F_sub  + n] <=  1e2)
    for n in range(Para.N_bus):
        model.addConstr(f_flow[N_F_load + n] ==  1)
    for n in range(Para.N_gen):
        model.addConstr(f_flow[N_F_gen  + n] == -1)

    # 2. Connectivity
    for n in range(Para.N_bus):
        # Bus-branch information
        line_head = Para.Line_head[n]
        line_tail = Para.Line_tail[n]
        # Formulate expression
        expr = gp.LinExpr()
        expr = expr - f_flow[N_F_load + n]
        expr = expr - gp.quicksum(f_flow[N_F_line + i] for i in line_head)
        expr = expr + gp.quicksum(f_flow[N_F_line + i] for i in line_tail)
        if n in Para.Sub[:,1]:
            i = int(np.where(n == Para.Sub[:,1])[0])
            expr = expr + f_flow[N_F_sub + i]
        if n in Para.Gen[:,1]:
            i = int(np.where(n == Para.Gen[:,1])[0])
            expr = expr + f_flow[N_F_gen + i]
        model.addConstr(expr == 0)
    
    # 3. Radial topology
    model.addConstr(gp.quicksum(y_line) == Para.N_bus - Para.N_sub)
    
    # Return
    return model


# This function defines the operation model
# -----------------------------------------------------------------------------
#
def DistFlow_Model(model, Para, y_line, v_flow, hour, obj):
    
    # Objective
    opr = gp.LinExpr()
    for n in range(Para.N_line):
        opr = opr + v_flow[N_I_line + n] * Para.Cost_los
    for n in range(Para.N_sub):  # power purchasing
        opr = opr + v_flow[N_P_sub  + n] * Para.Cost_sub
    for n in range(Para.N_bus):  # load shedding
        opr = opr + v_flow[N_P_cut  + n] * Para.Cost_pen
        opr = opr + v_flow[N_Q_cut  + n] * Para.Cost_pen
    for n in range(Para.N_gen):  # renewables
        opr = opr + v_flow[N_S_gen  + n] * Para.Gen[n,3]
    model.addConstr(obj[0] == opr)
    
    # Constraint
    # 1. Nodal active power balance
    for n in range(Para.N_bus):
        # Bus-Line information
        line_head = Para.Line_head[n]
        line_tail = Para.Line_tail[n]
        # Formulate expression
        expr = gp.LinExpr()
        expr = expr - gp.quicksum(v_flow[N_P_line + i] for i in line_head)
        expr = expr + gp.quicksum(v_flow[N_P_line + i] for i in line_tail)
        expr = expr + v_flow[N_P_cut + n]
        for i in line_tail:
            expr = expr - v_flow[N_I_line + i] * Para.Line[i,4]
        if n in Para.Sub[:,1]:  # active power input from substation
            i = int(np.where(n == Para.Sub[:,1])[0])
            expr = expr + v_flow[N_P_sub + i]
        if n in Para.Gen[:,1]:  # active power input from renewables
            i = int(np.where(n == Para.Gen[:,1])[0])
            expr = expr + v_flow[N_S_gen + i] * math.cos(Para.Factor)
        model.addConstr(expr == Para.Bus[n,1] * Para.Day[hour,1])
    
    # 2. Nodal reactive power balance
    for n in range(Para.N_bus):
        # Bus-Line information
        line_head = Para.Line_head[n]
        line_tail = Para.Line_tail[n]
        # Formulate expression
        expr = gp.LinExpr()
        expr = expr - gp.quicksum(v_flow[N_Q_line + i] for i in line_head)
        expr = expr + gp.quicksum(v_flow[N_Q_line + i] for i in line_tail)
        expr = expr + v_flow[N_Q_cut + n]
        for i in line_tail:
            expr = expr - v_flow[N_I_line + i] * Para.Line[i,5]
        if n in Para.Sub[:,1]:  # active power input from substation
            i = int(np.where(n == Para.Sub[:,1])[0])
            expr = expr + v_flow[N_Q_sub + i]
        if n in Para.Gen[:,1]:  # active power input from renewables
            i = int(np.where(n == Para.Gen[:,1])[0])
            expr = expr + v_flow[N_S_gen + i] * math.sin(Para.Factor)
        model.addConstr(expr == Para.Bus[n,2] * Para.Day[hour,1])
    
    # 3. Branch flow equation
    for n in range(Para.N_line):
        bus_head = Para.Line[n,1]
        bus_tail = Para.Line[n,2]
        # Formulate expression
        expr = gp.LinExpr()
        expr = expr + v_flow[N_V_bus + bus_head] - v_flow[N_V_bus + bus_tail]
        expr = expr - v_flow[N_P_line + n] * Para.Line[n,4] * 2
        expr = expr - v_flow[N_Q_line + n] * Para.Line[n,5] * 2
        expr = expr + v_flow[N_I_line + n] * Para.Line[n,4] ** 2 
        expr = expr + v_flow[N_I_line + n] * Para.Line[n,5] ** 2
        model.addConstr(expr >= -Para.Big_M * (1 - y_line[n]))
        model.addConstr(expr <=  Para.Big_M * (1 - y_line[n]))
    
    # 4. Second order conic constraint
    for n in range(Para.N_line):
        ep_0 = v_flow[N_P_line + n] * 2
        ep_1 = v_flow[N_Q_line + n] * 2
        ep_2 = v_flow[N_I_line + n] - v_flow[N_V_bus + Para.Line[n,1]]
        ep_3 = v_flow[N_I_line + n] + v_flow[N_V_bus + Para.Line[n,1]]
        model.addConstr(ep_0 * ep_0 + ep_1 * ep_1 + ep_2 * ep_2 <= ep_3 * ep_3)
    
    # 5. Renewables generation
    for n in range(Para.N_gen):
        expr = gp.LinExpr()
        expr = expr + v_flow[N_S_gen + n]
        expr = expr + v_flow[N_C_gen + n]
        G_type = int(Para.Gen[n,4])
        model.addConstr(expr == Para.Gen[n,2] * Para.Day[hour, G_type + 2])
    
    # 6. Lower and Upper bound
    # 1) voltage amplitutde
    for n in range(Para.N_bus):
        model.addConstr(v_flow[N_V_bus  + n] >= Para.V_min)
        model.addConstr(v_flow[N_V_bus  + n] <= Para.V_max)
    # 2) line current
    for n in range(Para.N_line):
        smax = (Para.Line[n,3] / Para.Base_V) ** 2
        model.addConstr(v_flow[N_I_line + n] >=  0)
        model.addConstr(v_flow[N_I_line + n] <=  y_line[n] * Para.Big_M)
        model.addConstr(v_flow[N_I_line + n] <=  smax)
    # 3) line flow
    for n in range(Para.N_line):
        smax = Para.Line[n,3]
        # active power
        model.addConstr(v_flow[N_P_line + n] >= -smax)
        model.addConstr(v_flow[N_P_line + n] <=  smax)
        model.addConstr(v_flow[N_P_line + n] >= -y_line[n] * Para.Big_M)
        model.addConstr(v_flow[N_P_line + n] <=  y_line[n] * Para.Big_M)
        # reactive power
        model.addConstr(v_flow[N_Q_line + n] >= -smax)
        model.addConstr(v_flow[N_Q_line + n] <=  smax)
        model.addConstr(v_flow[N_Q_line + n] >= -y_line[n] * Para.Big_M)
        model.addConstr(v_flow[N_Q_line + n] <=  y_line[n] * Para.Big_M)
    # 4) substation
    for n in range(Para.N_sub):
        smax = Para.Sub[n,2]
        model.addConstr(v_flow[N_P_sub  + n] >=  0)
        model.addConstr(v_flow[N_P_sub  + n] <=  smax)
        model.addConstr(v_flow[N_Q_sub  + n] >=  0)
        model.addConstr(v_flow[N_Q_sub  + n] <=  smax)
    # 5) renewables
    for n in range(Para.N_gen):
        G_type = int(Para.Gen[n,4])
        smax = Para.Gen[n,2] * Para.Day[hour, G_type + 2]
        model.addConstr(v_flow[N_S_gen  + n] >=  0)
        model.addConstr(v_flow[N_S_gen  + n] <=  smax)
        model.addConstr(v_flow[N_C_gen  + n] >=  0)
        model.addConstr(v_flow[N_C_gen  + n] <=  smax)
    # 6) load shedding
    for n in range(Para.N_bus):
        smax = Para.Bus[n,1]
        model.addConstr(v_flow[N_P_cut  + n] >=  0)
        model.addConstr(v_flow[N_P_cut  + n] <=  smax)
    for n in range(Para.N_bus):
        smax = Para.Bus[n,2]
        model.addConstr(v_flow[N_Q_cut  + n] >=  0)
        model.addConstr(v_flow[N_Q_cut  + n] <=  smax)

    # Return
    return model


if __name__ == "__main__":

    Para = Parameter("data/Data-IEEE-33.xlsx")
    tool = Excel_tool()
    hour = 12  # choose 12 a.m.

    # Optimize
    sol  = main_function(Para, hour)

    # Output
    var_list = [
        "y_line", "V_bus", "I_line", "P_line", "Q_line", "P_sub", "Q_sub", 
        "S_gen",  "C_gen", "P_cut",  "Q_cut"
    ]
    for item in var_list:
        print("--------------------------------------------------------------")
        print("The solution of {} is:".format(item))
        print(eval("sol.{}[:, 0]".format(item)))
        print("--------------------------------------------------------------")
    
    # Grapher
    fig  = Grapher(Para)
    fig.figure(Para, sol)

