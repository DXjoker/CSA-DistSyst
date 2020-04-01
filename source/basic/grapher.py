# !/usr/bin/python
#
# Copyright 2019, Southeast University, Liu Pengxiang
#
# IEEE PES-GM 2020 Paper
# A novel acceleration strategy for N-1 contingency screening in distribution system
#
# Grapher Toolbox


import numpy as np
import openpyxl as pyxl
import matplotlib.pyplot as plt


# This class creates the parameter class
#
class Parameter(object):

    # Initialization
    def __init__(self, filename):
        # System Data
        Data = self.ReadData(filename)
        self.Data  = Data[1]
        self.Load  = self.Data[:,1]
        self.Wind  = self.Data[:,2]
        self.Solar = self.Data[:,3]
        self.Water = self.Data[:,4]

    # This function inputs data from Excel files
    def ReadData(self, filename):
        data = []
        book = pyxl.load_workbook(filename)
        # Data preprocessing
        for i in book.sheetnames:  # sheet number
            sheet = book[i]
            n_row = sheet.max_row     # number of rows
            n_col = sheet.max_column  # number of columns
            data.append(self.Matrix_slice(sheet, n_row, n_col))  # append data
        return data

    # This function slices the matrix based on data
    def Matrix_slice(self, sheet, n_row, n_col):
        matrix = []
        k = 0
        for i in range(n_row):
            if sheet['A' + str(i + 1)].data_type == 'n':  # if value is a number
                matrix.append([])
                for j in range(n_col):
                    pos = chr(64 + j + 1) + str(i + 1)  # the position in the sheet
                    val = sheet[pos].value
                    matrix[k].append(val)
                k = k + 1
        return np.array(matrix)


def dayly_curve_divide(Para):

    plt.rcParams['savefig.dpi'] = 192
    plt.rcParams['figure.dpi']  = 192

    t = np.arange(0, 24, 0.25)
    fig, axs = plt.subplots(2, 2, figsize = (8, 8))
    # load
    y_min = min(Para.Load)
    y_max = max(Para.Load)
    axs[0,0].plot(t, Para.Load , color = '#4682B4')
    axs[0,0].set(xlim = (0, 23.75), ylim = (y_min - 0.025, y_max + 0.025))
    axs[0,0].fill_between(t, y_min - 0.025, Para.Load , color = '#4682B4', alpha = 0.3)
    axs[0,0].set_xlabel('Time / h \n (a) Load demand')
    axs[0,0].set_ylabel('load demand / p.u.')
    # wind
    y_min = min(Para.Wind)
    y_max = max(Para.Wind)
    axs[0,1].plot(t, Para.Wind , color = '#6B8E23')
    axs[0,1].set(xlim = (0, 23.75), ylim = (y_min - 0.025, y_max + 0.025))
    axs[0,1].fill_between(t, y_min - 0.025, Para.Wind , color = '#6B8E23', alpha = 0.3)
    axs[0,1].set_xlabel('Time / h \n (b) Wind farm')
    axs[0,1].set_ylabel('Generation / p.u.')
    # solar
    y_min = min(Para.Solar)
    y_max = max(Para.Solar)
    axs[1,0].plot(t, Para.Solar, color = '#D2691E')
    axs[1,0].set(xlim = (0, 23.75), ylim = (y_min - 0.002, y_max + 0.025))
    axs[1,0].fill_between(t, y_min - 0.002, Para.Solar, color = '#D2691E', alpha = 0.3)
    axs[1,0].set_xlabel('Time / h \n (c) PV station')
    axs[1,0].set_ylabel('Generation / p.u.')
    # water
    y_min = min(Para.Water)
    y_max = max(Para.Water)
    axs[1,1].plot(t, Para.Water, color = '#483D8B')
    axs[1,1].set(xlim = (0, 23.75), ylim = (y_min - 0.002, y_max + 0.025))
    axs[1,1].fill_between(t, y_min - 0.002, Para.Water, color = '#483D8B', alpha = 0.3)
    axs[1,1].set_xlabel('Time / h \n (d) Hydro power')
    axs[1,1].set_ylabel('Generation / p.u.')
    # plot
    plt.show()

def daily_curve_fusion(Para):
    # plt.rcParams['savefig.dpi'] = 192
    # plt.rcParams['figure.dpi']  = 192

    t = np.arange(0, 24, 0.25)
    width = 0.15
    p1 = plt.bar(t, Para.Wind , width)
    p2 = plt.bar(t, Para.Solar, width, bottom = Para.Wind)
    p3 = plt.bar(t, Para.Water, width, bottom = Para.Wind + Para.Solar)

    plt.show()


def curve_load(Para):
    plt.rcParams['figure.figsize'] = (3.24, 2)
    plt.rcParams['savefig.dpi'] = 192
    plt.rcParams['figure.dpi']  = 192

    t = np.arange(0, 24, 0.25)
    fig, ax = plt.subplots()
    # load
    y_min = min(Para.Load)
    y_max = max(Para.Load)

    ax.plot(t, Para.Load , color = '#4682B4')
    ax.set(xlim = (0, 23.75), ylim = (y_min - 0.025, y_max + 0.025))
    ax.fill_between(t, y_min - 0.025, Para.Load , color = '#4682B4', alpha = 0.3)
    ax.set_xlabel('Time / h \n (a) Load demand')
    ax.set_ylabel('load demand / p.u.')

    plt.show()


def curve_wind(Para):
    plt.rcParams['figure.figsize'] = (3.24, 2)
    plt.rcParams['savefig.dpi'] = 192
    plt.rcParams['figure.dpi']  = 192

    t = np.arange(0, 24, 0.25)
    fig, ax = plt.subplots()
    # Wind
    y_min = min(Para.Wind)
    y_max = max(Para.Wind)

    ax.plot(t, Para.Wind , color = '#FFA500')
    ax.set(xlim = (0, 23.75), ylim = (y_min - 0.025, y_max + 0.025))
    ax.fill_between(t, y_min - 0.025, Para.Wind , color = '#FFA500', alpha = 0.3)
    ax.set_xlabel('Time / h \n (b) Wind farm')
    ax.set_ylabel('Generation / p.u.')

    plt.show()


def curve_solar(Para):
    plt.rcParams['figure.figsize'] = (3.24, 2)
    plt.rcParams['savefig.dpi'] = 192
    plt.rcParams['figure.dpi']  = 192

    t = np.arange(0, 24, 0.25)
    fig, ax = plt.subplots()
    # Solar
    y_min = min(Para.Solar)
    y_max = max(Para.Solar)

    ax.plot(t, Para.Solar, color = '#6B8E23')
    ax.set(xlim = (0, 23.75), ylim = (y_min - 0.002, y_max + 0.025))
    ax.fill_between(t, y_min - 0.002, Para.Solar, color = '#6B8E23', alpha = 0.3)
    ax.set_xlabel('Time / h \n (c) PV station')
    ax.set_ylabel('Generation / p.u.')

    plt.show()
    

if __name__ == "__main__":

    # Input parameter
    Name = "data/Data-Curve.xlsx"  # file name
    Para = Parameter(Name)    # System parameter
    curve_solar(Para)

    n = 1