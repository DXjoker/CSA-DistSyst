# -----------------------------------------------------------------------------
# Copyright 2020, Southeast University, Liu Pengxiang
# 
# A novel acceleration strategy for N-1 contingency screening in distribution 
# system (data and callback driven)
# 
# -----------------------------------------------------------------------------


import sys
import math
import time
import numpy as np
import gurobipy as gp
import openpyxl as op
import matplotlib.pyplot as plt


# This class creates the System class
# -----------------------------------------------------------------------------
# 
class parameter(object):

    # Initialization
    def __init__(self, filename):

        # Inputs
        tool = excel_tool()  # excel file tool
        data = tool.read_excel(filename)
        
        # 1. Parameter of system
        # 1.1) line
        self.Line   = data[0]
        self.N_line = len(self.Line) # number of line
        # 1.2) bus
        self.Bus    = data[1]
        self.N_bus  = len(self.Bus)  # number of bus
        # 1.3) substation
        self.Sub    = data[2]
        self.N_sub  = len(self.Sub)  # number of substation
        # 1.4) reneables generation
        self.Gen    = data[3]
        self.N_gen  = len(self.Gen)  # number of renewables
        # 1.5) daily curve
        self.Day    = data[4]
        self.N_hour = len(self.Day)  # number of hours

        # 2. Parameter of base value
        # 2.1) base value
        self.Base_V = 12.66  # voltage: 12.66 kV
        self.Base_S = 10.00  # power:   10.00 MVA
        self.Base_Z = self.Base_V ** 2 / self.Base_S  # impedance
        self.Base_I = self.Base_S / self.Base_V / np.sqrt(3)  # current
        # 2.2) lower and upper limits
        self.V_min = (0.95 * self.Base_V) ** 2
        self.V_max = (1.05 * self.Base_V) ** 2

        # 3. Parameter of cost coefficient
        self.Cost_sub = 83   # cost of power purchasing
        self.Cost_pen = 200  # cost of load shedding
        self.Cost_los = 25   # cost of power loss

        # 4. Parameter of Bus-Line Information
        self.line_head = [[] for i in range(self.N_bus)]
        self.line_tail = [[] for i in range(self.N_bus)]
        for i in range(self.N_line):
            head = self.Line[i][1]
            tail = self.Line[i][2]
            self.line_head[int(round(head))].append(i)
            self.line_tail[int(round(tail))].append(i)
        
        # 5. Parameter of others
        self.Big_M  = 1e3  # a big number
        self.N_iter = 10   # number of iteration
        self.Factor = 0.31756  # power factor (rad)
    

# This class creates the index of variables
# -----------------------------------------------------------------------------
# 
class index_tool(object):
    
    # Initialization
    def __init__(self, para):

        # 1. Power flow (fictitious)
        self.N_F_line = 0                            # flow of line
        self.N_F_load = self.N_F_line + para.N_line  # flow of load demand
        self.N_F_sub  = self.N_F_load + para.N_bus   # flow of substation
        self.N_F_gen  = self.N_F_sub  + para.N_sub   # flow of DG
        self.N_F_var  = self.N_F_gen  + para.N_gen   # all fictitious variables
        
        # 2. Power flow (actual)
        self.N_V_bus  = 0                            # square of voltage
        self.N_I_line = self.N_V_bus  + para.N_bus   # square of current
        self.N_P_line = self.N_I_line + para.N_line  # line flow (P)
        self.N_Q_line = self.N_P_line + para.N_line  # line flow (Q)
        self.N_P_sub  = self.N_Q_line + para.N_line  # power injection (P)
        self.N_Q_sub  = self.N_P_sub  + para.N_sub   # power injection (Q)
        self.N_S_gen  = self.N_Q_sub  + para.N_sub   # renewables generation
        self.N_C_gen  = self.N_S_gen  + para.N_gen   # renewables curtailment
        self.N_P_cut  = self.N_C_gen  + para.N_gen   # Load shedding (P)
        self.N_Q_cut  = self.N_P_cut  + para.N_bus   # Load shedding (Q)
        self.N_V_var  = self.N_Q_cut  + para.N_bus   # all actual variables
    

# This class creates the result of the optimization
# -----------------------------------------------------------------------------
# 
class result(object):

    # Initialization
    def __init__(self):
        pass

    # Get value
    def getvalue(self, model, para, idx, f_flow, v_flow):

        # 1) power flow (fictitious)
        self.f_flow = f_flow.x
        self.f_line = self.f_flow[idx.N_F_line : idx.N_F_line + para.N_line]
        self.f_load = self.f_flow[idx.N_F_load : idx.N_F_load + para.N_bus ]
        self.f_sub  = self.f_flow[idx.N_F_sub  : idx.N_F_sub  + para.N_sub ]
        self.f_gen  = self.f_flow[idx.N_F_gen  : idx.N_F_gen  + para.N_gen ]
        # 2) power flow (actual)
        self.v_flow = v_flow.x
        self.V_bus  = self.v_flow[idx.N_V_bus  : idx.N_V_bus  + para.N_bus ]
        self.I_line = self.v_flow[idx.N_I_line : idx.N_I_line + para.N_line]
        self.P_line = self.v_flow[idx.N_P_line : idx.N_P_line + para.N_line]
        self.Q_line = self.v_flow[idx.N_Q_line : idx.N_Q_line + para.N_line]
        self.P_sub  = self.v_flow[idx.N_P_sub  : idx.N_P_sub  + para.N_sub ]
        self.Q_sub  = self.v_flow[idx.N_Q_sub  : idx.N_Q_sub  + para.N_sub ]
        self.S_gen  = self.v_flow[idx.N_S_gen  : idx.N_S_gen  + para.N_gen ]
        self.C_gen  = self.v_flow[idx.N_C_gen  : idx.N_C_gen  + para.N_gen ]
        self.P_cut  = self.v_flow[idx.N_P_cut  : idx.N_P_cut  + para.N_bus ]
        self.Q_cut  = self.v_flow[idx.N_Q_cut  : idx.N_Q_cut  + para.N_bus ]


# This class creates the execution tool for Excel files
# -----------------------------------------------------------------------------
# 
class excel_tool(object):

    # Initialization
    def __init__(self):
        pass

    # inputs data from Excel file
    def read_excel(self, filename):
        data = []
        book = op.load_workbook(filename)
        # data preprocessing
        for i, name in enumerate(book.sheetnames):  # sheet number
            if i < len(book.sheetnames):
                sheet = book[name]
                n_row = sheet.max_row     # number of rows
                n_col = sheet.max_column  # number of columns
                data.append(self.tool_filter(sheet, n_row, n_col))
        return data

    # saving data to excel file
    def save_excel(self, filename, sheetname, title, data):
        book  = op.load_workbook(filename)
        name  = book.sheetnames[-1]
        book.remove(book[name])
        sheet = book.create_sheet(sheetname)
        sheet.append(title)
        for i in range(len(data)):
            sheet.append(data[i,:].tolist())  # write data
        book.save(filename = filename)
    
    # filter data from the original numpy array
    def tool_filter(self, sheet, n_row, n_col):
        k = 0
        data = []
        for i in range(n_row):
            if sheet['A' + str(i + 1)].data_type == 'n':  # if it is a number
                data.append([])
                for j in range(n_col):
                    pos = chr(64 + j + 1) + str(i + 1)  # the position
                    val = sheet[pos].value
                    if sheet[pos].data_type == 'n':
                        data[k].append(val)
                k = k + 1
        return np.array(data)


# This class creates the drawing class
# -----------------------------------------------------------------------------
#
class grapher(object):

    # Initialization
    def __init__(self, para):
        # Switch line coordinate
        self.switch_line = []
        self.switch_line.append(np.array([[ 7, 0], [ 7,-2], [3 ,-2], [ 3,-3]]))
        self.switch_line.append(np.array([[ 8, 0], [ 8, 2], [14, 2], [14, 0]]))
        self.switch_line.append(np.array([[11, 0], [11,-3], [ 4,-3]]))
        self.switch_line.append(np.array([[17, 0], [17, 3], [12, 3]]))
        self.switch_line.append(np.array([[ 8, 6], [ 8, 5], [ 8, 3]]))
        # Line coordinate
        self.coordinate = []
        for n in range(para.N_line):
            if para.Line[n,6] == 0:  # line
                bus_head = int(round(para.Line[n,1]))
                bus_tail = int(round(para.Line[n,2]))
                x0 = para.Bus[bus_head, 3]
                y0 = para.Bus[bus_head, 4]
                x1 = para.Bus[bus_tail, 3]
                y1 = para.Bus[bus_tail, 4]
                self.coordinate.append(np.array([[x0,y0], [x1, y1]]))
            else:
                switch_no = int(para.Line[n,6] - 1)
                self.coordinate.append(self.switch_line[switch_no])
    
    # Figuring
    def figure(self, para, sol):

        for n in range(para.N_bus):
            plt.plot(para.Bus[n,3], para.Bus[n,4], 'b.')
            plt.text(para.Bus[n,3] + 0.05, para.Bus[n,4] + 0.10, '%s' % n)

        for n in range(para.N_line):
            for m in range(np.size(self.coordinate[n], 0) - 1):
                x0 = self.coordinate[n][m, 0]
                y0 = self.coordinate[n][m, 1]
                x1 = self.coordinate[n][m + 1, 0]
                y1 = self.coordinate[n][m + 1, 1]
                if sol.y_line[n] == 1:
                    plt.plot([x0, x1], [y0, y1], 'b-' )
                if sol.y_line[n] == 0:
                    plt.plot([x0, x1], [y0, y1], 'r--')
        plt.axis('equal')
        plt.show()


# This function filters the severe N-1 contingency
# -----------------------------------------------------------------------------
#
def contingency_screening(para, t, k):
    
    # 1. Pre-process
    # 1.1) input data
    idx = index_tool(para)
    # 1.2) build model
    model = gp.Model()

    # 2. Add variables
    i_sign = model.addMVar((1, 1), vtype = 'B')            # island sign
    x_line = model.addMVar((para.N_line, 3), vtype = 'B')  # configuration
    f_flow = model.addMVar((idx.N_F_var, 1), lb = -1e2)    # fictitious flow
    v_flow = model.addMVar((idx.N_V_var, 1), lb = -1e2)    # actual flow

    # 3. Add constraints
    model = add_islanding_constr(model, para, idx, x_line, i_sign)
    model = add_radiality_constr(model, para, idx, x_line, f_flow, i_sign)
    model = add_operating_constr(model, para, idx, x_line, v_flow, t)

    # 4. Add objective
    model = add_objective_expres(model, para, idx, v_flow)

    # 5. Set parameters
    # 5.1) lazy constraint callback
    for n in range(para.N_bus):
        model.addConstr(v_flow[idx.N_P_cut + n] == 0)
        model.addConstr(v_flow[idx.N_Q_cut + n] == 0)
    # 5.2) solver parameter
    model._x_line = x_line
    model._s_mild = np.zeros((para.N_line, 1))
    model.setParam("MIPGap", 0.001)
    model.Params.lazyConstraints = 1

    # 6. Solve
    # model.optimize()
    model.optimize(Lazy_Constraint_Callback)
    if model.status == gp.GRB.Status.OPTIMAL:
        # 1) solution
        sol = result()
        sol.getvalue(model, para, idx, f_flow, v_flow)
        sol.y_line = x_line[:, 0].x
        sol.c_line = x_line[:, 1].x
        sol.f_line = x_line[:, 2].x
        # 2) plot
        fig = grapher(para)
        fig.figure(para, sol)
    else:
        sol = model._s_mild
    return sol


# This function creates the islanding model
# -----------------------------------------------------------------------------
#
def add_islanding_constr(model, para, idx, x_line, i_sign):
    
    # 1. Pre-process
    y_line = x_line[:, 0]  # reconfiguration
    c_line = x_line[:, 1]  # contingency status
    f_line = x_line[:, 2]  # islanding mode

    # 2. Add constraints
    # 2.1) Topology
    for n in range(para.N_line):
        model.addConstr(c_line[n] - f_line[n] <= 1 - i_sign[0, 0])
        model.addConstr(y_line[n] + c_line[n] <= 1)
        model.addConstr(y_line[n] + f_line[n] <= 1)
    
    # 2.2) N-1 contingency and islanding
    model.addConstr(c_line.sum() == 1)
    model.addConstr(f_line.sum() == i_sign[0, 0])

    # 3. Return
    return model


# This function creates the radiality model
# -----------------------------------------------------------------------------
def add_radiality_constr(model, para, idx, x_line, f_flow, i_sign):
    
    # 1. Pre-process
    y_line = x_line[:, 0]  # reconfiguration
    c_line = x_line[:, 1]  # contingency status
    f_line = x_line[:, 2]  # islanding mode

    # 2 Add constraints
    # 2.1) fictitious power flow
    for n in range(para.N_line):
        expr = y_line[n] + f_line[n]
        model.addConstr(f_flow[idx.N_F_line + n] >= -para.Big_M * expr)
        model.addConstr(f_flow[idx.N_F_line + n] <=  para.Big_M * expr)
    for n in range(para.N_sub):
        model.addConstr(f_flow[idx.N_F_sub  + n] >=  0)
        model.addConstr(f_flow[idx.N_F_sub  + n] <=  para.Big_M)
    for n in range(para.N_bus):
        model.addConstr(f_flow[idx.N_F_load + n] ==  1)
    for n in range(para.N_gen):
        model.addConstr(f_flow[idx.N_F_gen  + n] == -1)
    
    # 2.2) connectivity
    for n in range(para.N_bus):
        # coefficient
        coef = np.zeros((1, idx.N_F_var))
        # bus-branch information
        line_head = para.line_head[n]
        line_tail = para.line_tail[n]
        # formulate coef-matrix
        coef[0, idx.N_F_load + n] = -1
        for i in line_head: coef[0, idx.N_F_line + i] = -1
        for i in line_tail: coef[0, idx.N_F_line + i] =  1
        if n in para.Sub[:,1]:
            i = int(np.where(n == para.Sub[:, 1])[0])
            coef[0, idx.N_F_sub + i] = 1
        if n in para.Gen[:,1]:
            i = int(np.where(n == para.Gen[:, 1])[0])
            coef[0, idx.N_F_gen + i] = 1
        # formulate constraint
        model.addConstr(coef @ f_flow[:, 0] == 0)

    # 2.3) radial topology
    model.addConstr(y_line.sum() == para.N_bus - para.N_sub - i_sign[0, 0])

    # 3. Return
    return model


# This function creates the operating model
# -----------------------------------------------------------------------------
#
def add_operating_constr(model, para, idx, x_line, v_flow, hour):
    
    # 1. Pre-process
    # 1.1) determine variables
    y_line = x_line[:, 0]  # reconfiguration
    c_line = x_line[:, 1]  # contingency status
    f_line = x_line[:, 2]  # islanding mode
    # 1.2) daily curve
    bus = np.zeros(para.N_bus)
    for n in range(para.N_bus):
        pick = 0.75 + np.random.rand(1) * 0.5
        bus[n] = para.Bus[n,1] * para.Day[hour,1] * pick
    gen = np.zeros(para.N_gen)
    for n in range(para.N_gen):
        pick = 0.75 + np.random.rand(1) * 0.5
        gen[n] = para.Gen[n,2] * para.Day[hour,int(para.Gen[n,4]) + 2] * pick

    # 2. Add constraints
    # 2.1) nodal power balance (P)
    for n in range(para.N_bus):
        # coefficient
        coef = np.zeros((1, idx.N_V_var))
        # bus-Line information
        line_head = para.line_head[n]
        line_tail = para.line_tail[n]
        # formulate coef-matrix
        coef[0, idx.N_P_cut + n] = 1
        for i in line_head: coef[0, idx.N_P_line + i] = -1
        for i in line_tail: coef[0, idx.N_P_line + i] =  1
        for i in line_tail:
            coef[0, idx.N_I_line + i] = -1 * para.Line[i, 4]
        if n in para.Sub[:, 1]:  # power input from substation
            i = int(np.where(n == para.Sub[:,1])[0])
            coef[0, idx.N_P_sub  + i] =  1
        if n in para.Gen[:, 1]:  # power input from renewables
            i = int(np.where(n == para.Gen[:,1])[0])
            coef[0, idx.N_S_gen  + i] =  1 * math.cos(para.Factor)
        # formulate constrant
        model.addConstr(coef @ v_flow[:, 0] == bus[n])
    
    # 2.2) nodal power balance (Q)
    for n in range(para.N_bus):
        # coefficient
        coef = np.zeros((1, idx.N_V_var))
        # bus-Line information
        line_head = para.line_head[n]
        line_tail = para.line_tail[n]
        # formulate coef-matrix
        coef[0, idx.N_Q_cut + n] = 1
        for i in line_head: coef[0, idx.N_Q_line + i] = -1
        for i in line_tail: coef[0, idx.N_Q_line + i] =  1
        for i in line_tail:
            coef[0, idx.N_I_line + i] = -1 * para.Line[i, 5]
        if n in para.Sub[:, 1]:  # power input from substation
            i = int(np.where(n == para.Sub[:,1])[0])
            coef[0, idx.N_Q_sub  + i] =  1
        if n in para.Gen[:, 1]:  # power input from renewables
            i = int(np.where(n == para.Gen[:,1])[0])
            coef[0, idx.N_S_gen  + i] =  1 * math.sin(para.Factor)
        # formulate constrant
        model.addConstr(coef @ v_flow[:, 0] == bus[n])
    
    # 2.3) branch flow equation
    for n in range(para.N_line):
        # coefficient
        coef = np.zeros((1, idx.N_V_var))
        # head and tail bus
        bus_head = int(para.Line[n, 1])
        bus_tail = int(para.Line[n, 2])
        # formulate coef-matrix
        coef[0, idx.N_V_bus + bus_head] =  1
        coef[0, idx.N_V_bus + bus_tail] = -1
        coef[0, idx.N_P_line + n] = -para.Line[n, 4] *  2
        coef[0, idx.N_Q_line + n] = -para.Line[n, 5] *  2
        coef[0, idx.N_I_line + n] =  para.Line[n, 4] ** 2 
        coef[0, idx.N_I_line + n] =  para.Line[n, 5] ** 2
        # formulate constrant
        model.addConstr(coef @ v_flow[:, 0] >= -para.Big_M * (1 - y_line[n]))
        model.addConstr(coef @ v_flow[:, 0] <=  para.Big_M * (1 - y_line[n]))
    
    # 2.4) renewables generation
    for n in range(para.N_gen):
        # coefficient
        coef = np.zeros((1, idx.N_V_var))
        # formulate coef-matrix
        coef[0, idx.N_S_gen + n] = 1
        coef[0, idx.N_C_gen + n] = 1
        # formulate constrant
        model.addConstr(coef @ v_flow[:, 0] == gen[n])
    
    # 2.5） linearise of second order conic constraints
    for n in range(para.N_line):
        # coefficient
        coef = np.zeros((1, idx.N_V_var))
        # formulate coef-matrix
        coef[0, idx.N_P_line + n] =  1
        coef[0, idx.N_Q_line + n] =  1
        expr = para.Line[n, 3]
        # formulate constrant
        model.addConstr(coef @ v_flow[:, 0] >= -1.414 * expr)
        model.addConstr(coef @ v_flow[:, 0] <=  1.414 * expr)
    for n in range(para.N_line):
        # coefficient
        coef = np.zeros((1, idx.N_V_var))
        # formulate coef-matrix
        coef[0, idx.N_P_line + n] =  1
        coef[0, idx.N_Q_line + n] = -1
        expr = para.Line[n, 3]
        # formulate constrant
        model.addConstr(coef @ v_flow[:, 0] >= -1.414 * expr)
        model.addConstr(coef @ v_flow[:, 0] <=  1.414 * expr)

    
    # 3. Add lower and upper bound
    # 3.1) voltage amplitude
    for n in range(para.N_bus):
        model.addConstr(v_flow[idx.N_V_bus  + n] >= para.V_min)
        model.addConstr(v_flow[idx.N_V_bus  + n] <= para.V_max)
    # 3.2) line current
    for n in range(para.N_line):
        imax = (para.Line[n,3] / para.Base_V) ** 2
        model.addConstr(v_flow[idx.N_I_line + n] >=  0)
        model.addConstr(v_flow[idx.N_I_line + n] <=  y_line[n] * para.Big_M)
        model.addConstr(v_flow[idx.N_I_line + n] <=  imax)
    # 3.3) line flow
    for n in range(para.N_line):
        smax = para.Line[n,3]
        # active power
        model.addConstr(v_flow[idx.N_P_line + n] >= -smax)
        model.addConstr(v_flow[idx.N_P_line + n] <=  smax)
        model.addConstr(v_flow[idx.N_P_line + n] >= -y_line[n] * para.Big_M)
        model.addConstr(v_flow[idx.N_P_line + n] <=  y_line[n] * para.Big_M)
        # reactive power
        model.addConstr(v_flow[idx.N_Q_line + n] >= -smax)
        model.addConstr(v_flow[idx.N_Q_line + n] <=  smax)
        model.addConstr(v_flow[idx.N_Q_line + n] >= -y_line[n] * para.Big_M)
        model.addConstr(v_flow[idx.N_Q_line + n] <=  y_line[n] * para.Big_M)
    # 3.4) substation
    for n in range(para.N_sub):
        smax = para.Sub[n,2]
        model.addConstr(v_flow[idx.N_P_sub  + n] >=  0)
        model.addConstr(v_flow[idx.N_P_sub  + n] <=  smax)
        model.addConstr(v_flow[idx.N_Q_sub  + n] >=  0)
        model.addConstr(v_flow[idx.N_Q_sub  + n] <=  smax)
    # 3.5) renewables
    for n in range(para.N_gen):
        model.addConstr(v_flow[idx.N_S_gen  + n] >=  0)
        model.addConstr(v_flow[idx.N_S_gen  + n] <=  gen[n])
        model.addConstr(v_flow[idx.N_C_gen  + n] >=  0)
        model.addConstr(v_flow[idx.N_C_gen  + n] <=  gen[n])
    # 3.6) load shedding
    for n in range(para.N_bus):
        smax = para.Bus[n,1]
        model.addConstr(v_flow[idx.N_P_cut  + n] >=  0)
        model.addConstr(v_flow[idx.N_P_cut  + n] <=  smax)
    for n in range(para.N_bus):
        smax = para.Bus[n,2]
        model.addConstr(v_flow[idx.N_Q_cut  + n] >=  0)
        model.addConstr(v_flow[idx.N_Q_cut  + n] <=  smax)
    
    # Return
    return model


# This function creates the objective expression
# -----------------------------------------------------------------------------
# 
def add_objective_expres(model, para, idx, v_flow):
    
    # 1. Formulate coefficient
    coef = np.zeros((1, idx.N_V_var))
    # 1.1) power loss
    for n in range(para.N_line):
        coef[0, idx.N_I_line + n] = para.Cost_los
    # 1.2) power purchasing
    for n in range(para.N_sub):
        coef[0, idx.N_P_sub  + n] = para.Cost_sub
    # 1.3) load shedding
    for n in range(para.N_bus):
        coef[0, idx.N_P_cut  + n] = para.Cost_pen
        coef[0, idx.N_Q_cut  + n] = para.Cost_pen
    # 1.4) renewable generation
    for n in range(para.N_gen):
        coef[0, idx.N_S_gen  + n] = para.Gen[n, 3]
    
    # 2. Add objective
    model.setObjective(coef @ v_flow[:, 0], gp.GRB.MINIMIZE)

    # Return
    return model


# This function creates lazy constraints callback approach
# -----------------------------------------------------------------------------
#
def Lazy_Constraint_Callback(model, where):

    # Incumbent solution
    if where == gp.GRB.Callback.MIPSOL:

        # 1. get incumbent solution
        x_line = model.cbGetSolution(model._x_line)
        y_line = np.round(x_line[:, 0])
        c_line = np.round(x_line[:, 1])

        # 2. add constraint
        pos = np.where(c_line == 1)
        model._s_mild[pos, 0] = 1
        model.cbLazy(model._x_line[pos, 1].vararr[0,0] == 0)


# Main function
# -----------------------------------------------------------------------------
# 
if __name__ == "__main__":

    # Parameter
    filename = "data/Data-IEEE-33.xlsx"
    para = parameter(filename)
    tool = excel_tool()

    mat = np.zeros((para.N_line, para.N_line))

    for t in range(para.N_hour):
        for k in range(para.N_iter):
            sol = contingency_screening(para, t, k)
            for i in np.where(sol == 1)[0]:
                for j in np.where(sol == 1)[0]:
                    mat[i, j] = mat[i, j] + 1
    
    name = "result/matrix.xlsx"
    tool.save_excel(name, "result", [" res "], mat)